import pandas as pd
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template import loader
from rest_framework.decorators import authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db import connection
from pmc_api.custom_permissions import *
import calendar
from datetime import datetime
from pmc_api.serializers import ApplicantDetailSerializer
from pmc_api.models import ApplicantDetail
from openpyxl import Workbook
# For formatting (openpyxl)
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import json

class ReportAPIView(APIView):
    
    permission_classes = [IsAuthenticated, IsInAnalytics2Group]
    def get(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        query1 = """
                SELECT 
                d.district_name, 
                count(CASE WHEN a.registration_for = 'Producer' THEN b.id else NULL END) as Producer,
                count(CASE WHEN a.registration_for = 'Consumer' THEN b.id else NULL END) as Consumer,
                count(CASE WHEN a.registration_for = 'Collector' THEN b.id else NULL END) as Collector,
                count(CASE WHEN a.registration_for = 'Recycler' THEN b.id else NULL END) as Recycler,
                count(b.id) as Total
                FROM 
                tbl_districts d
                LEFT JOIN pmc_api_businessprofile g on g.district_id = d.district_id
                LEFT JOIN public.pmc_api_applicantdetail a on a.id = g.applicant_id
                LEFT JOIN 
                    pmc_api_applicationsubmitted b ON a.id = b.applicant_id
                LEFT JOIN pmc_api_applicantfee f on a.id = f.applicant_id
                GROUP BY 1
                -- HAVING count(b.id) > 1
                ORDER BY 1;
        """
        # Second query
        query2 = """
                    SELECT DISTINCT ON (a.tracking_number) 
                    substring(a.tracking_number, 1, 3) AS district,
                    a.tracking_number,
                    (b.created_at)::timestamp AS received_on,
                    (c.created_at)::timestamp AS LSO_ASSIGNED_AT
                FROM public.pmc_api_applicantdetail a
                JOIN pmc_api_applicationsubmitted b ON a.id = b.applicant_id
                JOIN pmc_api_applicationassignment c ON a.id = c.applicant_id
                WHERE a.assigned_group = 'DO'
                  AND c.assigned_group = 'DO'
                ORDER BY a.tracking_number, b.created_at;
                """
        
        query3 = """
        SELECT 
            td.short_name AS district_short_name, 
            count(DISTINCT a.tracking_number) AS count_do_applications_in,
            count(distinct c2a.tracking_number) AS count_do_applications_out_forward,
            count(distinct c1a.tracking_number) AS count_do_applications_out_backward,
            count(distinct c2a.tracking_number) + count(distinct c1a.tracking_number) as count_do_applications_out_total
        FROM
            public.pmc_api_applicantdetail a
        JOIN 
            pmc_api_applicationassignment c ON a.id = c.applicant_id
            and c.assigned_group = 'DO'
        JOIN pmc_api_businessprofile bf on c.id = bf.applicant_id
        JOIN 
            pmc_api_applicationsubmitted b ON a.id = b.applicant_id
        FULL OUTER JOIN tbl_districts td ON bf.district_id = td.district_id
        LEFT JOIN pmc_api_applicationassignment c2 ON a.id = c2.applicant_id 
            and c2.assigned_group = 'LSM2'
            LEFT JOIN public.pmc_api_applicantdetail c2a ON c2.applicant_id = c2a.id
        LEFT JOIN pmc_api_applicationassignment c1 ON a.id = c1.applicant_id 
            and c1.assigned_group = 'LSM' and c1.id > c.id
            LEFT JOIN public.pmc_api_applicantdetail c1a ON c1.applicant_id = c1a.id
        GROUP BY 
            1
        ORDER BY 
            1, 2
        """
        
        # Fee Report Query
        query4 = """
            SELECT 
                d.district_name,
                c.created_at::date,
                sum(a.fee_amount) AS fee_generated_submitted,
                -- sum(case when c.id is not null then a.fee_amount end) as fee_submitted,
                sum(case when a.is_settled then a.fee_amount end) as fee_verified
                FROM 
                pmc_api_applicantfee a
                JOIN pmc_api_applicantdetail b ON a.applicant_id = b.id
                JOIN pmc_api_businessprofile p ON a.applicant_id = p.applicant_id
                JOIN pmc_api_applicationsubmitted c on a.applicant_id = c.applicant_id
                FULL OUTER JOIN tbl_districts d on p.district_id = d.district_id
                GROUP BY 
                    GROUPING SETS (
                        (d.district_name, c.created_at::date),  -- Group by district and date
                        (d.district_name)  -- Rollup row for the district only (without date)
                    )
                ORDER BY 
                    d.district_name, c.created_at::date
        """

        # === NEW QUERIES ===
        query5 = """
            -- Fee Report - 1 Daily
            SELECT
                COALESCE(d.district_name, 'Total') AS district_name,
                sum(a.fee_amount) AS fee_submitted,
                sum(case when a.is_settled then a.fee_amount end) as fee_verified,
                sum(a.fee_amount) - COALESCE(sum(case when a.is_settled then a.fee_amount else null end)::numeric,0) as fee_unverified
            FROM 
                pmc_api_applicantfee a
            JOIN pmc_api_applicantdetail b ON a.applicant_id = b.id
            JOIN pmc_api_businessprofile p ON a.applicant_id = p.applicant_id
            JOIN pmc_api_applicationsubmitted c on a.applicant_id = c.applicant_id
            JOIN tbl_districts d on p.district_id = d.district_id
            WHERE c.created_at < '2025-01-01'
            GROUP BY 
                GROUPING SETS (
                    (d.district_name),  -- Group by district and date
                    ()  -- Rollup row for the district only (without date)
                )
            ORDER BY 
                d.district_name
        """

        query6 = """
            -- Fee Report - 2 Daily
            SELECT
                COALESCE(d.district_name, 'Total') AS district_name,
                sum(a.fee_amount) AS fee_submitted,
                sum(case when a.is_settled then a.fee_amount end) as fee_verified,
                sum(a.fee_amount) - COALESCE(sum(case when a.is_settled then a.fee_amount else null end)::numeric,0) as fee_unverified
            FROM 
                pmc_api_applicantfee a
            JOIN pmc_api_applicantdetail b ON a.applicant_id = b.id
            JOIN pmc_api_businessprofile p ON a.applicant_id = p.applicant_id
            JOIN pmc_api_applicationsubmitted c on a.applicant_id = c.applicant_id
            JOIN tbl_districts d on p.district_id = d.district_id
            WHERE c.created_at BETWEEN '2025-01-01' AND '2025-01-09'
            GROUP BY 
                GROUPING SETS (
                    (d.district_name),  -- Group by district and date
                    ()  -- Rollup row for the district only (without date)
                )
            ORDER BY 
                d.district_name
        """

        try:
            with connection.cursor() as cursor:
                # Execute first query
                cursor.execute(query1)
                columns1 = [col[0] for col in cursor.description]
                results1 = cursor.fetchall()

                # Execute second query
                cursor.execute(query2)
                columns2 = [col[0] for col in cursor.description]
                results2 = cursor.fetchall()

                # Execute third query
                cursor.execute(query3)
                columns3 = [col[0] for col in cursor.description]
                results3 = cursor.fetchall()
                
                # Check if user has 'Analytics3' group or permission
                user_has_permission = request.user.groups.filter(name='Analytics3').exists() or request.user.has_perm('IsInAnalytics3Group')

                # Execute query for fee report if the user meets the condition
                if user_has_permission:
                    cursor.execute(query4)
                    columns4 = [col[0] for col in cursor.description]
                    results4 = cursor.fetchall()
                    
                    # === Execute new queries for two additional fee reports ===
                    cursor.execute(query5)
                    columns5 = [col[0] for col in cursor.description]
                    results5 = cursor.fetchall()

                    cursor.execute(query6)
                    columns6 = [col[0] for col in cursor.description]
                    results6 = cursor.fetchall()

        except Exception as e:
            return Response({"error": str(e)}, status=500)

        # Convert the results to DataFrames
        df1 = pd.DataFrame(results1, columns=columns1)
        df2 = pd.DataFrame(results2, columns=columns2)
        df3 = pd.DataFrame(results3, columns=columns3)

        # Create an HTTP response with an Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        # === Utility function to format new Fee sheets ===
        def format_fee_sheet(sheet, title_text):
            """
            1) Insert a row at the top for a merged title from A1 to D1.
            2) Bold + center the title.
            3) Format columns B, C, D as #,##0.00
            4) Widen columns A..D to width=20
            """
            from openpyxl.styles import Font, Alignment

            # Insert a blank row at top so existing data moves down
            sheet.insert_rows(1)

            # Put title in A1, merge A1:D1
            sheet["A1"] = title_text
            sheet.merge_cells("A1:D1")
            sheet["A1"].font = Font(bold=True)
            sheet["A1"].alignment = Alignment(horizontal='center')

            # Format columns 2..4 (B..D) with thousands separator, 2 decimals
            for col_letter in ['B', 'C', 'D']:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet[f"{col_letter}{row}"]
                    cell.number_format = '#,##0.00'

            # Widen columns A..D
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, 5):
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = 20

            # === Bold the last row (columns A..D) ===
            last_row_idx = sheet.max_row
            for col_idx in range(1, 5):
                col_letter = get_column_letter(col_idx)
                last_cell = sheet[f"{col_letter}{last_row_idx}"]
                last_cell.font = Font(bold=True)
                
        # Write DataFrames to separate sheets in the Excel file
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df1.to_excel(writer, index=False, sheet_name='Summary Report')
            df2.to_excel(writer, index=False, sheet_name='DO Report')
            df3.to_excel(writer, index=False, sheet_name='DO Application Summary Report')
            
            # If the user has Analytics3 group or authority, add df4 to the Excel
            if user_has_permission:
                df4 = pd.DataFrame(results4, columns=columns4)
                df4.to_excel(writer, index=False, sheet_name='Fee Report')
                
                # For the new queries
                df5 = pd.DataFrame(results5, columns=columns5)
                df6 = pd.DataFrame(results6, columns=columns6)
                
                df5['fee_submitted'] = df5['fee_submitted'].astype(float)
                df5['fee_verified']           = df5['fee_verified'].astype(float)
                df5['fee_unverified']         = df5['fee_unverified'].astype(float)

                df6['fee_submitted'] = df6['fee_submitted'].astype(float)
                df6['fee_verified']           = df6['fee_verified'].astype(float)
                df6['fee_unverified']         = df6['fee_unverified'].astype(float)
                
                # 1) Fee Daily 1 (query5)
                df5.to_excel(writer, index=False, sheet_name='Fee Daily 1')
                sheet5 = writer.book['Fee Daily 1']
                format_fee_sheet(sheet5, "18.12.24 to 31.12.25")

                # 2) Fee Daily 2 (query6)
                df6.to_excel(writer, index=False, sheet_name='Fee Daily 2')
                sheet6 = writer.book['Fee Daily 2']
                format_fee_sheet(sheet6, "01.01.25 to 08.01.25")

        return response

def generate_cutoff_dates(start_year=2025, start_month=2):
    # Start with given cutoff dates
    cutoff_dates = ['2024-12-31', '2025-01-08', '2025-01-31']

    # Get the current year and month dynamically
    today = datetime.today()
    current_year, current_month, current_day = datetime.today().year, datetime.today().month, datetime.today().day


    # Start from February 2025
    year, month = 2025, 2

    while (year, month) <= (current_year, current_month):  # Stop at current month and year
        # Get last day of the month, considering the current month condition
        if year == current_year and month == current_month:
            last_day = current_day  # Use today's date for the current month
        else:
            last_day = calendar.monthrange(year, month)[1]  # Use month's last day for past months

        cutoff_date = f"{year}-{month:02d}-{last_day}"
        cutoff_dates.append(cutoff_date)

        # Move to the next month
        if month == 12:
            month = 1
            year += 1
        else:
            month += 1

    return cutoff_dates


class ReportFeeAPIView(APIView):
    
    permission_classes = [IsAuthenticated, IsInAnalytics3Group]

    def get(self, request):
        """
        Fetch fee statistics for specific cutoff dates.
        """
        # Generate cutoff dates up to 2026
        cutoff_dates = generate_cutoff_dates()
        results = []

        previous_cutoff_date = None

        for date in cutoff_dates:
            with connection.cursor() as cursor:
                if previous_cutoff_date is None:
                    # First iteration: Only apply upper limit
                    cursor.execute("""
                        SELECT
                            %s as till,  -- Date cutoff
                            sum(a.fee_amount) AS fee_received,
                            sum(case when a.is_settled then a.fee_amount end) as fee_verified,
                            sum(a.fee_amount) - COALESCE(sum(case when a.is_settled then a.fee_amount else null end)::numeric, 0) as fee_unverified
                        FROM 
                            pmc_api_applicantfee a
                        JOIN pmc_api_applicantdetail b ON a.applicant_id = b.id
                        JOIN pmc_api_businessprofile p ON a.applicant_id = p.applicant_id
                        JOIN pmc_api_applicationsubmitted c on a.applicant_id = c.applicant_id
                        JOIN tbl_districts d on p.district_id = d.district_id
                        WHERE c.created_at::DATE <= %s
                        AND b.assigned_group != 'APPLICANT'
                    """, [date, date])
                else:
                    # Subsequent iterations: Apply both lower and upper limits
                    cursor.execute("""
                        SELECT
                            %s as till,  -- Date cutoff
                            sum(a.fee_amount) AS fee_received,
                            sum(case when a.is_settled then a.fee_amount end) as fee_verified,
                            sum(a.fee_amount) - COALESCE(sum(case when a.is_settled then a.fee_amount else null end)::numeric, 0) as fee_unverified
                        FROM 
                            pmc_api_applicantfee a
                        JOIN pmc_api_applicantdetail b ON a.applicant_id = b.id
                        JOIN pmc_api_businessprofile p ON a.applicant_id = p.applicant_id
                        JOIN pmc_api_applicationsubmitted c on a.applicant_id = c.applicant_id
                        JOIN tbl_districts d on p.district_id = d.district_id
                        WHERE c.created_at::DATE > %s AND c.created_at::DATE <= %s
                        AND b.assigned_group != 'APPLICANT'
                    """, [date, previous_cutoff_date, date])

                row = cursor.fetchone()
                if row:
                    results.append({
                        "till": row[0],
                        "fee_received": row[1] or 0,
                        "fee_verified": row[2] or 0,
                        "fee_unverified": row[3] or 0
                    })

            previous_cutoff_date = date  # Update previous cutoff date

        return Response(results)


class ExportApplicantDetailsToExcelAPIView(APIView):
    
    permission_classes = [IsAuthenticated]
    # permission_classes = []

    def post(self, request):
            
        if request.method != "POST":
                return HttpResponse("Only POST method is allowed", status=405)

        try:
            body = json.loads(request.body)
            applicant_ids = body.get("applicant_ids", [])

            if not applicant_ids:
                return HttpResponse("No applicant IDs provided.", status=400)

            # Fetch only selected applicants
            applicants = ApplicantDetail.objects.filter(id__in=applicant_ids)
        
        except json.JSONDecodeError:
            return HttpResponse("Invalid JSON format", status=400)
        
        # Serialize data
        serializer = ApplicantDetailSerializer(applicants, many=True)
        data = serializer.data

        # Define column groupings
        column_groups = {
            "Applicant Information": ["tracking_number", "first_name", "last_name", "gender", "cnic", "email", "mobile_no", "application_status"],
            "Business Information": ["businessprofile.business_name", "businessprofile.entity_type", "businessprofile.district.district_name", "businessprofile.tehsil.tehsil_name", "businessprofile.city_town_village", "businessprofile.postal_address"],
            "Misc. Details": ["registration_for", "has_identity_document", "has_fee_challan", "assigned_group"],
            "Producer Details": ["producer.tracking_number", "producer.registration_required_for", "producer.number_of_machines", "producer.total_capacity_value", "producer.date_of_setting_up", "producer.total_waste_generated_value", "producer.has_waste_storage_capacity", "producer.waste_disposal_provision"],
            "Stockist/Distributor/Supplier Details": ["consumer.registration_required_for", "consumer.consumption", "consumer.provision_waste_disposal_bins", "consumer.no_of_waste_disposable_bins", "consumer.segregated_plastics_handed_over_to_registered_recyclers"],
            "Collector Details": ["collector.registration_required_for", "collector.selected_categories", "collector.total_capacity_value", "collector.number_of_vehicles", "collector.number_of_persons"],
            "Recycler Details": ["recycler.selected_categories", "recycler.plastic_waste_acquired_through", "recycler.has_adequate_pollution_control_systems", "recycler.pollution_control_details"],
            "Financial Details": ["total_fee_amount", "verified_fee_amount", "psid_tracking.consumer_number", "psid_tracking.payment_status"],
            "Manual Fields": ["manual_fields.latitude", "manual_fields.longitude", "manual_fields.list_of_products", "manual_fields.list_of_by_products", "manual_fields.raw_material_imported", "manual_fields.seller_name_if_raw_material_bought", "manual_fields.self_import_details", "manual_fields.raw_material_utilized", "manual_fields.compliance_thickness_75", "manual_fields.valid_consent_permit_building_bylaws", "manual_fields.stockist_distributor_list", "manual_fields.procurement_per_day", "manual_fields.no_of_workers", "manual_fields.labor_dept_registration_status", "manual_fields.occupational_safety_and_health_facilities", "manual_fields.adverse_environmental_impacts"]
        }
        
        # Color mapping for sections
        group_colors = {
            "Applicant Information": "FFA500",
            "Business Information": "009000",
            "Misc. Details": "0000FF",
            "Producer Details": "FF4500",
            "Stockist/Distributor/Supplier Details": "1E90FF",
            "Recycler Details": "32CD32",
            "Collector Details": "FFD700",
            "Financial Details": "8A2BE2",
            "Manual Fields": "A9A9A9"
        }
        
        # Create an Excel workbook
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Applicant_Details_{datetime.now()}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Applicant Details"
        
        # Add Report Header
        ws.merge_cells("A1:N1")
        header_cell = ws["A1"]
        header_cell.value = "Plastic Management Information System - Applicant Details"
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A2:N2")
        date_cell = ws["A2"]
        date_cell.value = f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(bold=True, size=12)
        date_cell.alignment = Alignment(horizontal="center")
        
        # Prepare headers with groupings
        headers = []
        group_headers = []
        
        for group, columns in column_groups.items():
            headers.extend(columns)
            group_headers.extend([group] * len(columns))
        
        # Write group headers with background color
        ws.append(group_headers)
        
        col_start = 1
        for group, columns in column_groups.items():
            col_end = col_start + len(columns) - 1
            ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
            group_cell = ws.cell(row=3, column=col_start)
            group_cell.value = group
            group_cell.font = Font(bold=True)
            group_cell.alignment = Alignment(horizontal="center")
            group_cell.fill = PatternFill(start_color=group_colors.get(group, "FFFFFF"), end_color=group_colors.get(group, "FFFFFF"), fill_type="solid")
            col_start = col_end + 1
        
        # Write column headers
        ws.append(headers)
        
        # Apply styles to headers
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 20  # Adjust column width
        
        # Append data
        for applicant in data:
            row = []
            registration_for = applicant.get("registration_for", "")
            for group, columns in column_groups.items():
                for column in columns:
                    value = applicant
                    for key in column.split("."):
                        value = value.get(key, "") if isinstance(value, dict) else ""
                    
                    # Convert lists to comma-separated strings
                    if isinstance(value, list):
                        value = ", ".join(map(str, value))
                    
                    # Leave Producer, Consumer, Recycler, Collector empty if registration_for does not match
                    if group in ["Producer Details", "Stockist/Distributor/Supplier Details", "Recycler Details", "Collector Details"] and registration_for not in ["Producer", "Consumer", "Recycler", "Collector"]:
                        value = ""
                    
                    row.append(value)
            ws.append(row)
        
        # Save the workbook to the response
        wb.save(response)
        return response


class ApplicantFeeReportExcel(APIView):

    permission_classes = [IsAuthenticated, IsInAnalytics2Group]

    def get(self, request):
        query = """
        SELECT
            a.id,
            d.district_name,
            a.tracking_number,
            a.first_name,
            a.last_name,
            g.name as business_name,
            CAST(b.created_at AS TIMESTAMP) submission_date_time,
            (case when b.id is not null then f.fee_amount else null end) as fee_submitted,
            pt.consumer_number as consumer_number_psid,
            pt.bank_code,
            pt.paid_date,
            pt.paid_time,
            pt.amount_bifurcation,
            pt.dept_transaction_id
        FROM 
            public.pmc_api_applicantdetail a
        LEFT JOIN pmc_api_applicationsubmitted b ON a.id = b.applicant_id
        LEFT JOIN pmc_api_applicantfee f on a.id = f.applicant_id
        LEFT JOIN pmc_api_businessprofile g on a.id = g.applicant_id
        LEFT JOIN tbl_districts d on g.district_id = d.district_id
        JOIN pmc_api_psidtracking pt ON a.id = pt.applicant_id and pt.payment_status = 'PAID'
        WHERE
            (case when b.id is not null then f.fee_amount else null end) is not null
            AND a.assigned_group in ('LSM', 'DO', 'LSM2')
            AND exists (
                SELECT 1 FROM pmc_api_psidtracking t
                WHERE t.applicant_id = a.id AND t.payment_status = 'PAID'
            )
            AND NOT EXISTS (SELECT 1 FROM pmc_api_applicantdocuments b
				WHERE b.applicant_id = a.id
				AND document_description = 'Fee Verification from Treasury/District Accounts Office'
			)
        ORDER BY pt.paid_date, pt.paid_time
        """

        try:
            with connection.cursor() as cursor:
                cursor.execute(query)
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()

        except Exception as e:
            return HttpResponse(f"Error: {str(e)}", status=500)

        df = pd.DataFrame(results, columns=columns)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="applicant_fee_report.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Applicant Fee Report')

            workbook = writer.book
            worksheet = writer.sheets['Applicant Fee Report']

            # Set column widths and bold headers
            for col_cells in worksheet.iter_cols(min_row=1, max_row=1):
                col_letter = col_cells[0].column_letter
                worksheet.column_dimensions[col_letter].width = 25
                for cell in col_cells:
                    cell.font = Font(bold=True)

        return response