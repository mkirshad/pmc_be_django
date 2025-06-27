from rest_framework import viewsets, permissions, status
from rest_framework.response import Response

from .models import *
from .serializers import *
from rest_framework.decorators import action
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.shortcuts import get_object_or_404
import pdfkit
from django.contrib.auth.decorators import login_required
from .custom_permissions import *
from django.db.models import Q, Count, When, Case, Sum, Value, IntegerField, F, Func, IntegerField, Exists, OuterRef
from django.core.exceptions import ValidationError
import os
from collections import OrderedDict
from django.conf import settings
from pmc_api.threadlocals import get_current_user

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

from django.template.loader import render_to_string
from rest_framework.parsers import MultiPartParser, FormParser
from django.views.decorators.http import require_GET

def log_access(request, model_name, object_id):
    return
    AccessLog.objects.create(
        user=get_current_user(),
        model_name=model_name,
        object_id=str(object_id),
        method=request.method,
        ip_address=request.META.get('REMOTE_ADDR'),
        endpoint=request.path
    )

def get_plastic_types(applicant_detail):
    reg_for = (applicant_detail.registration_for or "").strip()
    default_plastics = "N/A"

    try:
        if reg_for == 'Producer':
            producer = Producer.objects.get(applicant=applicant_detail)
            reg_list = producer.registration_required_for or []
            reg_other_list = producer.registration_required_for_other or []
            final_list = reg_list + reg_other_list
        elif reg_for == 'Consumer':
            consumer = Consumer.objects.get(applicant=applicant_detail)
            reg_list = consumer.registration_required_for or []
            reg_other_list = consumer.registration_required_for_other or []
            final_list = reg_list + reg_other_list
        elif reg_for == 'Collector':
            collector = Collector.objects.get(applicant=applicant_detail)
            reg_list = collector.registration_required_for or []
            reg_other_list = collector.registration_required_for_other or []
            final_list = reg_list + reg_other_list
        elif reg_for == 'Recycler':
            recycler = Recycler.objects.get(applicant=applicant_detail)
            final_list = recycler.selected_categories or []
        else:
            return default_plastics

        # Convert dicts to strings if needed
        cleaned = []
        for item in final_list:
            if isinstance(item, dict):
                cleaned.append(", ".join(f"{k}: {v}" for k, v in item.items()))
            else:
                cleaned.append(str(item))

        return ", ".join(cleaned) if cleaned else default_plastics

    except (Producer.DoesNotExist,
            Consumer.DoesNotExist,
            Collector.DoesNotExist,
            Recycler.DoesNotExist):
        return default_plastics

def get_particulars(applicant_detail):
    """
    Returns the string for the License.particulars field:
      - If registration_for is 'Producer', include the number_of_machines from Producer model.
      - Otherwise, read entity_type from BusinessProfile and indicate 'Business Type: X'.
    """

    reg_for = (applicant_detail.registration_for or "").lower().strip()

    # Default if no data or no matching sub-model
    default_particulars = "N/A"

    if reg_for == "producer":
        # Try to fetch Producer model
        from .models import Producer
        try:
            producer = Producer.objects.get(applicant=applicant_detail)
            num_machines = producer.number_of_machines or "Unknown"
            return f"Number of Machines: {num_machines}"
        except Producer.DoesNotExist:
            return default_particulars
    else:
        # For Consumer, Collector, Recycler, or any other type
        # Check if applicant_detail has a BusinessProfile:
        business_profile = getattr(applicant_detail, 'businessprofile', None)
        if business_profile:
            entity_type = business_profile.entity_type or "Unknown"
            return f"Business Type: {entity_type}"
        return default_particulars

# Create or update License
def create_or_update_license(applicant_detail, user=None):
    """
    Helper function to either create or update a License record 
    whenever assigned_group = 'Download License'.
    
    - applicant_detail: an ApplicantDetail instance
    - user: the user performing the update (optional, used for created_by if needed)
    """
    # print(applicant_detail.assigned_group)
    # 1) Basic check: Only proceed if assigned_group == 'Download License'
    if applicant_detail.assigned_group != 'Download License':
        return  # No license creation if not in correct group

    # 2) Gather or compute data needed for the license
    #    Example logic (customize to your domain rules):
    license_for = applicant_detail.registration_for

    # Owner name: combine first/last
    owner_name = f"{applicant_detail.first_name} {applicant_detail.last_name}".strip()

    # Business name & address from applicant_detail.businessprofile, if it exists
    business_name = ""
    address = ""
    if hasattr(applicant_detail, 'businessprofile') and applicant_detail.businessprofile:
        bp = applicant_detail.businessprofile
        business_name = bp.business_name or bp.name or ""
        address = bp.postal_address or ""

    # Suppose you want to sum all fees (whether settled or not). 
    # If you want only settled fees, filter(is_settled=True)
    fee_amount = ApplicantFee.objects.filter(applicant=applicant_detail).aggregate(
        total=Sum('fee_amount')
    )['total'] or 0

    # Types of plastics logic. For now, just a placeholder. 
    # You could gather from applicant_detail or from a Producer/Consumer model.
    types_of_plastics = get_plastic_types(applicant_detail)
    
        # **New**: Get particulars for Producer or others
    particulars_str = get_particulars(applicant_detail)

    # License number logic. Maybe we use the applicant_detail's tracking_number 
    # or fallback to a pattern:
    license_number = applicant_detail.tracking_number

    # License duration. This can be dynamic, we just use "3 Years" as an example.
    license_duration = "3 Years"

    # The date_of_issue can be "today" or a custom logic
    date_of_issue = timezone.now().date()

    max_len = 200
    license_data = {
        "license_for": license_for,
        "license_number": license_number,
        "license_duration": license_duration,
        "owner_name": owner_name[:max_len],
        "business_name": business_name[:max_len],
        "types_of_plastics": types_of_plastics[:max_len],
        "particulars": particulars_str[:max_len],
        "fee_amount": fee_amount,
        "address": address[:300],  # assuming address has 300 max_length
        "date_of_issue": date_of_issue,
        "created_by": user,
    }
    # 3) Create or update the license record (we assume one license per applicant per date, or just 1 record total)
    #    If you only ever want one license total per applicant, you could ignore date_of_issue in the lookup
    obj, created = License.objects.update_or_create(
        applicant_id=applicant_detail.pk,
        defaults=license_data
    )
    # End of function. The License record now exists or has been updated.


class ApplicantDetailViewSet(viewsets.ModelViewSet):
    queryset = ApplicantDetail.objects.all()

    serializer_class = ApplicantDetailSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        """
        Optionally restricts the returned queryset to records created by the currently authenticated user.
        If all records should be accessible, remove this method.
        """
        user = self.request.user
        target_groups = {'LSO', 'LSM', 'DO', 'TL', 'MO', 'LSM2','DEO', 'DG', 'Download License'}
        user_groups = set(user.groups.values_list('name', flat=True))
        matching_groups = user_groups.intersection(target_groups)
        print(matching_groups)
        # Allow 'Super' group to access all records
        if "Super" in user_groups:
            print("Super user matched - returning all records")
            queryset = super().get_queryset()

            # Filter by assigned_group
            assigned_group = self.request.query_params.get("assigned_group")
            if assigned_group:
                queryset = queryset.filter(assigned_group=assigned_group)

            # Filter by application_status
            application_status = self.request.query_params.get("application_status")
            if application_status:
                queryset = queryset.filter(application_status=application_status)

            return queryset

        
        
        # Check for LSO specific filtering
        if "LSO" in matching_groups and user.username.lower().startswith("lso."):
            print('LSO is matched')
            username = user.username.lower()  # Ensure case-insensitivity
            if username.startswith("lso."):
                try:
                    user_suffix = int(username.split(".")[1])  # Extract the numeric suffix
                except (IndexError, ValueError):
                    return ApplicantDetail.objects.none()  # Return empty queryset for invalid usernames

                # Fetch IDs matching the pattern based on user's suffix
                ids_to_include = ApplicationSubmitted.objects.annotate(
                    mod_index=(models.F("id") - user_suffix) % 3  # Calculate modulo based on user suffix
                ).filter(mod_index=0).values_list("applicant_id", flat=True)

                return ApplicantDetail.objects.filter(
                    assigned_group__in=["LSO", "APPLICANT"],
                    id__in=ids_to_include,
                )
            
        # Check for DO specific filtering
        elif "DO" in matching_groups and user.username.lower().startswith("do."):
            print('do is matched')
            username = user.username.lower()
            if username.startswith("do."):
                try:
                    district_code = username.split(".")[1].upper()  # Extract district code from username and convert to uppercase
                except IndexError:
                    return ApplicantDetail.objects.none()  # Return empty queryset for invalid usernames
                print('going to return DO records')
                return ApplicantDetail.objects.filter(
                    assigned_group="DO",
                    businessprofile__district__short_name=district_code  # Match district code
                )
        elif matching_groups:
            return ApplicantDetail.objects.filter(assigned_group__in=matching_groups)
        else:
            return ApplicantDetail.objects.filter(created_by=user)
  

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    

    def destroy(self, request, *args, **kwargs):
        """
        Restrict delete operation: only allow superusers to delete records.
        """
        applicant_detail = self.get_object()  # Get the object to be deleted

        # Check if the user is a superuser
        if not request.user.is_superuser:
            return Response(
                {'detail': 'You do not have permission to delete this record.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Proceed with the delete if the user is a superuser
        return super().destroy(request, *args, **kwargs)

    def perform_update(self, serializer):
        # Remove 'tracking_hash' from validated data if present
        print('its in update of applicantDetail')
        validated_data = serializer.validated_data
        validated_data.pop('tracking_hash', None)  # Prevent tracking_hash updates
        instance = serializer.save()
        # Check assigned_group and trigger license creation
        create_or_update_license(instance, user=self.request.user)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()

        # ✅ Log access when record is retrieved
        log_access(
            request=request,
            model_name='ApplicantDetail',
            object_id=instance.id
        )

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def list(self, request, *args, **kwargs):
        log_access(
            request=request,
            model_name='ApplicantDetail',
            object_id='LIST'
        )
        return super().list(request, *args, **kwargs)

class ApplicantDetailMainListViewSet(viewsets.ModelViewSet):
    queryset = ApplicantDetail.objects.all()

    serializer_class = ApplicantDetailMainListSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        """
        Optionally restricts the returned queryset to records created by the currently authenticated user.
        If all records should be accessible, remove this method.
        """
        user = self.request.user
        target_groups = {'LSO', 'LSM', 'DO', 'TL', 'MO', 'LSM2','DEO','DG', 'Download License'}
        user_groups = set(user.groups.values_list('name', flat=True))
        matching_groups = user_groups.intersection(target_groups)
        print(matching_groups)
        # Allow 'Super' group to access all records
        if "Super" in user_groups:
            print("Super user matched - returning all records")
            queryset = super().get_queryset()

            # Filter by assigned_group
            assigned_group = self.request.query_params.get("assigned_group")
            if assigned_group:
                if(assigned_group == 'Submitted'):
                    # Only include records with a related ApplicationSubmitted record
                    queryset = queryset.filter(
                        Exists(ApplicationSubmitted.objects.filter(applicant=OuterRef('pk')))
                    )
                elif assigned_group == 'PMC':
                    # Include all records where assigned_group is LSO, LSM, LSM2, or TL
                    queryset = queryset.filter(assigned_group__in=['LSO', 'LSM', 'LSM2', 'TL'])
                else:
                    queryset = queryset.filter(assigned_group=assigned_group)

            # Filter by application_status
            application_status = self.request.query_params.get("application_status")
            if application_status:
                queryset = queryset.filter(application_status=application_status)
                
                

            return queryset
        
class ApplicantDetailMainDOListViewSet(viewsets.ModelViewSet):
    """
    A ViewSet for listing ApplicantDetail records filtered by user's district and optional query parameters.
    """
    queryset = ApplicantDetail.objects.all()
    serializer_class = ApplicantDetailMainListSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        """
        Restricts the queryset to records for the district associated with the logged-in user,
        and optionally filters by assigned_group or application_status.
        """
        user = self.request.user
        target_groups = {'LSO', 'LSM', 'DO', 'TL', 'MO', 'LSM2', 'DEO', 'DG', 'Download License'}
        user_groups = set(user.groups.values_list('name', flat=True))

        # Extract district short_name from username (e.g., "do.lhr")
        username = user.username
        district_short_name = username.split('.')[1].upper() if '.' in username else None

        # Allow 'Super' group to access all records
        if "DO" not in user_groups:
            print("Super user matched - returning all records")
            raise ValueError("Not DO Group")
        else:
            # Ensure district short_name is valid
            if not district_short_name:
                raise ValueError("Invalid username format. District short name missing.")

            # Fetch the district based on short_name (case-insensitive)
            try:
                district = TblDistricts.objects.get(short_name__iexact=district_short_name)
            except TblDistricts.DoesNotExist:
                raise ValueError("No matching district found for the user.")

            # Filter queryset by the user's district
            queryset = super().get_queryset().filter(businessprofile__district=district)

        # Apply additional filters
        assigned_group = self.request.query_params.get("assigned_group")
        if assigned_group:
            queryset = queryset.filter(assigned_group=assigned_group)

        application_status = self.request.query_params.get("application_status")
        if application_status:
            queryset = queryset.filter(application_status=application_status)

        # Print the SQL query
        sql_query = str(queryset.query)
        print(f"Generated SQL Query:\n{sql_query}")
        
        return queryset

class ApplicantManualFieldsViewSet(viewsets.ModelViewSet):
    queryset = ApplicantManualFields.objects.all()
    serializer_class = ApplicantManualFieldsSerializer

class ApplicantFieldResponseViewSet(viewsets.ModelViewSet):
    queryset = ApplicantFieldResponse.objects.all()
    serializer_class = ApplicantFieldResponseSerializer
    permission_classes = [permissions.IsAuthenticated]

    def create(self, request, *args, **kwargs):
        # Check if the request data is a list
        is_bulk = isinstance(request.data, list)

        if is_bulk:
            # Use many=True to handle bulk data
            serializer = self.get_serializer(data=request.data, many=True)
        else:
            serializer = self.get_serializer(data=request.data)

        serializer.is_valid(raise_exception=True)
        # Save data with the current user
        if is_bulk:
            serializer.save(created_by=self.request.user)
        else:
            serializer.save(created_by=self.request.user)

        # Return the appropriate response
        if is_bulk:
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        # Ensure created_by is always set
        serializer.save(created_by=self.request.user)

class BusinessProfileViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing, creating, updating, and deleting BusinessProfile records.
    """
    serializer_class = BusinessProfileSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        """
        Returns BusinessProfile instances associated with the authenticated user.
        """
        return BusinessProfile.objects.all()  # Replace with filtering logic if needed

    def perform_create(self, serializer):
        """
        Automatically set updated_by during creation.
        """
        serializer.save(updated_by=self.request.user)
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        """
        Automatically update updated_by when a record is updated.
        """
        serializer.save(updated_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        """
        Custom delete behavior, if needed.
        """
        instance = self.get_object()
        if not request.user.is_superuser:
            return Response(
                {'detail': 'Only superusers can delete this record.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'])
    def by_applicant(self, request):
        applicant_id = request.query_params.get('applicant_id')
        if applicant_id:
            profiles = self.get_queryset().filter(applicant_id=applicant_id)
            serializer = self.get_serializer(profiles, many=True)
            return Response(serializer.data)
        return Response({'error': 'applicant_id is required'}, status=400)


class PlasticItemsViewSet(viewsets.ModelViewSet):
    queryset = PlasticItems.objects.all()
    serializer_class = PlasticItemsSerializer
    permission_classes = [permissions.IsAuthenticated]


class ProductsViewSet(viewsets.ModelViewSet):
    queryset = Products.objects.all()
    serializer_class = ProductsSerializer
    permission_classes = [permissions.IsAuthenticated]


class ByProductsViewSet(viewsets.ModelViewSet):
    queryset = ByProducts.objects.all()
    serializer_class = ByProductsSerializer
    permission_classes = [permissions.IsAuthenticated]


class ProducerViewSet(viewsets.ModelViewSet):
    queryset = Producer.objects.all()
    serializer_class = ProducerSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def create(self, request, *args, **kwargs):
        applicant_id = request.data.get('applicant')

        if not applicant_id:
            return Response({"error": "Applicant is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if a Producer already exists for this applicant
        try:
            producer = Producer.objects.get(applicant_id=applicant_id, created_by=self.request.user.id)
            serializer = self.get_serializer(producer, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save(created_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Producer.DoesNotExist:
            # If no producer exists, proceed with the normal creation flow
            return super().create(request, *args, **kwargs)
        
class ConsumerViewSet(viewsets.ModelViewSet):
    queryset = Consumer.objects.all()
    serializer_class = ConsumerSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def create(self, request, *args, **kwargs):
        applicant_id = request.data.get('applicant')

        if not applicant_id:
            return Response({"error": "Applicant is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if a Producer already exists for this applicant
        try:
            consumer = Consumer.objects.get(applicant_id=applicant_id, created_by=self.request.user.id)
            serializer = self.get_serializer(consumer, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save(created_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Consumer.DoesNotExist:
            # If no producer exists, proceed with the normal creation flow
            return super().create(request, *args, **kwargs)
        
class CollectorViewSet(viewsets.ModelViewSet):
    queryset = Collector.objects.all()
    serializer_class = CollectorSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def create(self, request, *args, **kwargs):
        applicant_id = request.data.get('applicant')

        if not applicant_id:
            return Response({"error": "Applicant is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if a Producer already exists for this applicant
        try:
            consumer = Collector.objects.get(applicant_id=applicant_id, created_by=self.request.user.id)
            serializer = self.get_serializer(consumer, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save(created_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Collector.DoesNotExist:
            # If no producer exists, proceed with the normal creation flow
            return super().create(request, *args, **kwargs)

class RecyclerViewSet(viewsets.ModelViewSet):
    queryset = Recycler.objects.all()
    serializer_class = RecyclerSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def create(self, request, *args, **kwargs):
        applicant_id = request.data.get('applicant')

        if not applicant_id:
            return Response({"error": "Applicant is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if a Producer already exists for this applicant
        try:
            recycler = Recycler.objects.get(applicant_id=applicant_id, created_by=self.request.user.id)
            serializer = self.get_serializer(recycler, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save(created_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Recycler.DoesNotExist:
            # If no producer exists, proceed with the normal creation flow
            return super().create(request, *args, **kwargs)
        
class RawMaterialViewSet(viewsets.ModelViewSet):
    queryset = RawMaterial.objects.all()
    serializer_class = RawMaterialSerializer
    permission_classes = [permissions.IsAuthenticated]

# class DivisionViewSet(viewsets.ReadOnlyModelViewSet):
#     """
#     API endpoint that allows divisions to be viewed.
#     """
#     queryset = TblDivisions.objects.all()
#     serializer_class = DivisionSerializer

class DistrictViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint that allows districts to be viewed.
    """
    permission_classes = [IsAuthenticated]
    queryset = TblDistricts.objects.all()
    serializer_class = DistrictSerializer

class DistrictGEOMViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint that allows districts to be viewed.
    """
    permission_classes = []
    queryset = TblDistricts.objects.exclude(geom=None)
    serializer_class = DistrictGEOMSerializer
    
class TehsilViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint that allows tehsils to be viewed.
    """
    queryset = TblTehsils.objects.all()
    serializer_class = TehsilSerializer

    def get_queryset(self):
        """
        Optionally restrict the queryset to tehsils of a specific district.
        """
        queryset = super().get_queryset()
        district_id = self.request.query_params.get('district_id')  # Get district_id from query params
        if district_id:
            queryset = queryset.filter(district_id=district_id)  # Filter by district_id
        return queryset

class DistrictByLatLonSet(APIView):
    """
    API endpoint that returns the district name based on latitude and longitude.
    """

    def get(self, request):
        """
        Retrieve the district name based on latitude and longitude.
        """
        lat = request.query_params.get('lat')
        lon = request.query_params.get('lon')

        # Validate input
        if not lat or not lon:
            return Response(
                {"error": "Latitude and longitude are required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            lat, lon = float(lat), float(lon)  # Convert to float
        except ValueError:
            return Response(
                {"error": "Invalid latitude or longitude format."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Fetch district
        district = TblDistricts.get_district_by_coordinates(lat, lon)
        
        if district:
            return Response({"district_name": district}, status=status.HTTP_200_OK)
        else:
            return Response(
                {"error": "District not found for the given coordinates."},
                status=status.HTTP_404_NOT_FOUND
            )
class UserGroupsView(generics.ListAPIView):
    serializer_class = GroupSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.request.user.groups.all()

class UserGroupsViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = GroupSerializer

    def get_queryset(self):
        return self.request.user.groups.all()

    def get_serializer_context(self):
        """
        Pass request context to serializer to access user details
        """
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
class ApplicationAssignmentViewSet(viewsets.ModelViewSet):
    
    
    # queryset = ApplicationAssignment.objects.all().order_by('-created_at')
    serializer_class = ApplicationAssignmentSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        return ApplicationAssignment.objects.all().order_by('-created_at')
    
    @action(detail=False, methods=['get'])
    def by_applicant(self, request):
        applicant_id = request.query_params.get('applicant_id')
        if applicant_id:
            profiles = self.get_queryset().filter(applicant_id=applicant_id)
            serializer = self.get_serializer(profiles, many=True)
            return Response(serializer.data)
        return Response({'error': 'applicant_id is required'}, status=400)

    def perform_create(self, serializer):
    # Ensure the applicant is correctly passed to the serializer
        print('its in Application Assignment!')
        # print(instance.assigned_group)
        applicant_id = self.request.data.get('applicant')
        if not applicant_id:
            raise ValidationError({"applicant": "Applicant is required."})

        try:
            # Fetch the ApplicantDetail instance
            applicant = ApplicantDetail.objects.get(id=applicant_id)
        except ApplicantDetail.DoesNotExist:
            raise ValidationError({"applicant": "Invalid applicant ID."})

        # Save the ApplicationAssignment instance
        instance = serializer.save(created_by=self.request.user, applicant=applicant)

        # Update the assigned_group in the related ApplicantDetail
        print(instance.assigned_group)
        applicant.assigned_group = instance.assigned_group
        applicant.application_status = 'In Process'
        applicant.save()
        
        create_or_update_license(applicant, user=self.request.user)

def get_original_host(request):
    # Prefer HTTP_ORIGIN as it is more reliable for CORS requests
    origin = request.META.get('HTTP_ORIGIN', '')

    # Fallback to HTTP_REFERER if HTTP_ORIGIN is not available
    if not origin:
        referer = request.META.get('HTTP_REFERER', '')
        if referer:
            # Extract scheme + host from the referer URL
            from urllib.parse import urlparse
            parsed_url = urlparse(referer)
            origin = f"{parsed_url.scheme}://{parsed_url.netloc}"
    return origin

class ApplicantDocumentsViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    queryset = ApplicantDocuments.objects.all()
    serializer_class = ApplicanDocumentsSerializer
    def get_serializer_context(self):
        # Get the original host and pass it to the serializer context
        context = super().get_serializer_context()
        context['original_host'] = get_original_host(self.request)
        return context

    def create(self, request, *args, **kwargs):
        document_description = request.data.get("document_description")
        applicant_id = request.data.get("applicant")
        document = request.data.get("document")

        if not applicant_id or not document_description or not document:
            return Response({'error': 'Missing required fields'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Save the document
            applicant = get_object_or_404(ApplicantDetail, id=applicant_id)
            document_instance = ApplicantDocuments.objects.create(
                applicant=applicant,
                document=document,
                document_description=document_description,
                created_by=request.user,
            )

            # Check if the document description matches the fee verification
            if document_description == "Fee Verification from Treasury/District Accounts Office":
                unsettled_fee = ApplicantFee.objects.filter(applicant=applicant, is_settled=False).first()
                if unsettled_fee:
                    unsettled_fee.is_settled = True
                    unsettled_fee.save()
                    
            # Serialize the saved document instance
            serializer = self.get_serializer(document_instance)
            return Response(serializer.data, status=status.HTTP_201_CREATED)


        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class Modulo(Func):
    """
    Custom function to perform modulo operation in Django ORM.
    """
    function = "MOD"
    template = "%(function)s(%(expressions)s)"
    output_field = IntegerField()

class FetchStatisticsViewSet(viewsets.ModelViewSet):
    """
    A ViewSet for fetching application statistics grouped by assigned_group.
    """

    def list(self, request):
        # Get all groups assigned to the user
        user_groups = request.user.groups.all()
        
        # If the user is not assigned to any group, return an empty response
        if not user_groups.exists():
            return Response({})
        
        # Extract the actual group names of the user
        assigned_group_names = [group.name for group in user_groups]

        # Check if the user is in 'Super' group
        is_super_user = "Super" in assigned_group_names
        
        # If you only want to filter by user groups (except for Super special handling),
        # you can do something like:
        allowed_groups = [group[0] for group in USER_GROUPS]

        # Fetch counts for assigned groups
        statistics = (
            ApplicantDetail.objects
            .filter(assigned_group__in=allowed_groups)
            .values('assigned_group')
            .annotate(count=Count('id'))
        )

        # Create a dictionary that initializes counts for "All-Applications" and "Challan-Downloaded"
        # plus each group the user has
        if is_super_user:
            result = {
                group: 0
                for group in ["All-Applications", "Challan-Downloaded", "Submitted"] + ["PMC"]+ allowed_groups
            }
            # "All-Applications" -> total number of applicants
            result["All-Applications"] = ApplicantDetail.objects.count()
            # "Challan-Downloaded" -> total with application_status != "Created"
            result["Challan-Downloaded"] = ApplicantDetail.objects.filter(
                application_status="Fee Challan"
            ).count()
            result["Submitted"] = ApplicationSubmitted.objects.count()
            # If the group is LSO, add LSO1, LSO2, and LSO3 right after
            lso_submitted = ApplicationSubmitted.objects.annotate(
                mod_value=Modulo(F('id'), 3)
            )

            lso1_ids = lso_submitted.filter(mod_value=1).values_list('applicant_id', flat=True)
            lso2_ids = lso_submitted.filter(mod_value=2).values_list('applicant_id', flat=True)
            lso3_ids = lso_submitted.filter(mod_value=0).values_list('applicant_id', flat=True)

            result["LSO1"] = ApplicantDetail.objects.filter(assigned_group = 'LSO', id__in=lso1_ids).count()
            result["LSO2"] = ApplicantDetail.objects.filter(assigned_group = 'LSO', id__in=lso2_ids).count()
            result["LSO3"] = ApplicantDetail.objects.filter(assigned_group = 'LSO', id__in=lso3_ids).count()

        else:
            result = {
                group: 0
                for group in allowed_groups
            }

        # Fill in actual counts for each group

        
         
        for stat in statistics:
            # Calculate 'PMC' as sum of 'LSO', 'LSM', 'LSM2', 'TL'
            result[stat['assigned_group']] = stat['count']

        result["PMC"] = (
                result.get("LSO", 0) + 
                result.get("LSM", 0) + 
                result.get("LSM2", 0) + 
                result.get("TL", 0)
            )
        

        # Otherwise, return everything
        return Response(result)

class FetchStatisticsDOViewSet(viewsets.ModelViewSet):
    """
    A ViewSet for fetching application statistics grouped by assigned_group.
    """

    def list(self, request):
        # Get all groups assigned to the user
        user_groups = request.user.groups.all()

        # If the user is not assigned to any group, return an empty response
        if not user_groups.exists():
            return Response({})

        # Extract the actual group names of the user
        assigned_group_names = [group.name for group in user_groups]

        # Check if the user is in 'Super' group
        is_super_user = "Super" in assigned_group_names

        # If you only want to filter by user groups (except for Super special handling),
        # you can do something like:
        allowed_groups = [group[0] for group in USER_GROUPS_DO]

        # Extract district short_name from the username (e.g., "do.lhr")
        username = request.user.username
        district_short_name = username.split('.')[1].upper() if '.' in username else None  # Convert to uppercase

        # Ensure short_name is valid
        if not district_short_name:
            return Response({"error": "Invalid username format. District short name missing."}, status=400)

        # Fetch the district based on short_name (case-insensitive)
        try:
            district = TblDistricts.objects.get(short_name__iexact=district_short_name)
        except TblDistricts.DoesNotExist:
            return Response({"error": "No matching district found for the user."}, status=400)

        # Filter ApplicantDetail by district and exclude records with empty or NULL assigned_group
        statistics = (
            ApplicantDetail.objects
            .filter(businessprofile__district=district)
            .exclude(assigned_group__isnull=True)  # Exclude NULL values
            .exclude(assigned_group="")  # Exclude empty strings
            .values('assigned_group')
            .annotate(count=Count('id'))
        )

        # Create a dictionary that initializes counts for "allowed_groups"
        result = {group: 0 for group in allowed_groups}

        # Fill in actual counts for each group
        for stat in statistics:
            result[stat['assigned_group']] = stat['count']

        # Return the resulting dictionary
        return Response(result)

def generate_license_pdf(request):
    # Get parameters from the request
    applicant_id = request.GET.get("applicant_id")
    tracking_number = request.GET.get("tracking_number")

    # Fetch ApplicantDetail based on the given parameter
    if applicant_id:
        applicant = get_object_or_404(ApplicantDetail, id=applicant_id)
    elif tracking_number:
        applicant = get_object_or_404(ApplicantDetail, tracking_number=tracking_number)
    else:
        return JsonResponse({"error": "Either 'applicant_id' or 'tracking_number' must be provided."}, status=400)

    # Fetch BusinessProfile linked to the ApplicantDetail
    business_profile = applicant.businessprofile if hasattr(applicant, 'businessprofile') else None

    # Define data for the PDF
    pdf_data = {
        "license_number": applicant.tracking_number,
        "license_duration": "1 Year",  # You can customize this value
        "owner_name": f"{applicant.first_name} {applicant.last_name or ''}",
        "business_name": business_profile.name if business_profile else "N/A",
        "address": business_profile.postal_address if business_profile else "N/A",
        "cnic_number": applicant.cnic,
    }

    # Create the HTML content for the PDF
    html_content = f"""
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
            }}
            .header {{
                text-align: center;
                font-size: 24px;
                margin-bottom: 20px;
            }}
            .content {{
                margin: 0 20px;
                line-height: 1.5;
            }}
            .content div {{
                margin: 10px 0;
            }}
        </style>
    </head>
    <body>
        <div class="header">Plastic License Certificate</div>
        <div class="content">
            <div><strong>License Number:</strong> {pdf_data['license_number']}</div>
            <div><strong>License Duration:</strong> {pdf_data['license_duration']}</div>
            <div><strong>Owner Name:</strong> {pdf_data['owner_name']}</div>
            <div><strong>Business Name:</strong> {pdf_data['business_name']}</div>
            <div><strong>Address:</strong> {pdf_data['address']}</div>
            <div><strong>CNIC Number:</strong> {pdf_data['cnic_number']}</div>
        </div>
    </body>
    </html>
    """

    # Generate the PDF using pdfkit
    try:
        pdf = pdfkit.from_string(html_content, False)  # Generate PDF as binary data
    except Exception as e:
        return JsonResponse({"error": f"PDF generation failed: {str(e)}"}, status=500)

    # Return the PDF as a response
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f"inline; filename=license_{applicant.tracking_number}.pdf"
    return response

# @login_required
def download_latest_document(request):
    # Get the logged-in user
    user = request.user

    # Validate the applicant associated with the logged-in user
    applicant = get_object_or_404(ApplicantDetail, created_by=user.id)

    # Get document description from request parameters
    document_description = request.GET.get("document_description")
    if not document_description:
        return JsonResponse({"error": "Document description parameter is required."}, status=400)

    # Query the latest document matching the applicant and document description
    try:
        document = (
            ApplicantDocuments.objects.filter(
                applicant=applicant,
                document_description=document_description
            )
            .order_by("-created_at")  # Latest document first
            .first()
        )

        if not document:
            return JsonResponse(
                {"error": f"No document found with description '{document_description}'."},
                status=404
            )

        # Prepare the response for file download
        response = HttpResponse(document.document.open(), content_type="application/octet-stream")
        response["Content-Disposition"] = f"attachment; filename={document.document.name.split('/')[-1]}"
        return response

    except Exception as e:
        return JsonResponse({"error": f"An error occurred: {str(e)}"}, status=500)

# BI Statistics and Analytics
class ApplicantStatisticsView(APIView):
    """
    API to calculate applicant statistics and return details for the grid.
    """
    permission_classes = [IsAuthenticated, IsInAnalytics1Group]
    def get(self, request):
        # Aggregated counts for all applicants
        # applications_count = ApplicantDetail.objects.exclude(application_status='Created').count()
        # licenses_count = ApplicantDetail.objects.filter(application_status='Completed').count()
        # in_progress_count = ApplicantDetail.objects.filter(
        #     ~Q(application_status='Created') & 
        #     ~Q(application_status='Completed') & 
        #     ~Q(application_status='Rejected')
        # ).count()
        # # New
        # do_count = ApplicantDetail.objects.filter(
        #     ~Q(application_status='Created') & 
        #     ~Q(application_status='Completed') & 
        #     ~Q(application_status='Rejected') &
        #     Q(assigned_group='DO')
        # ).count()
        # pmc_count = ApplicantDetail.objects.filter(
        #     ~Q(application_status='Created') & 
        #     ~Q(application_status='Completed') & 
        #     ~Q(application_status='Rejected') &
        #     ~Q(assigned_group='DO')
        # ).count()
        # fee_challan_count = ApplicantDetail.objects.filter(application_status='Fee Challan').count()
        # in_progress_count = ApplicantDetail.objects.filter(
        #     ~Q(application_status='Created') & 
        #     ~Q(application_status='Completed') & 
        #     ~Q(application_status='Rejected')
        # ).count()
        # in_progress_count = ApplicantDetail.objects.filter(
        #     ~Q(application_status='Created') & 
        #     ~Q(application_status='Completed') & 
        #     ~Q(application_status='Rejected')
        # ).count()

        # District-wise statistics grouped by 'registration_for'
        district_statistics = ApplicantDetail.objects.exclude(
            application_status__in=['Created', 'Fee Challan']
        ) \
            .values('registration_for', 'businessprofile__district__district_name') \
            .annotate(count=Count('id')) \
            .order_by('businessprofile__district__district_name')

        
        # Statistics filtered by 'registration_for'
        registration_statistics = ApplicantDetail.objects \
            .exclude(registration_for__isnull=True) \
            .values('registration_for') \
            .annotate(
                Applications=Count('id', filter=~Q(application_status='Created') &
                                        ~Q(application_status='Fee Challan')),
                DO=Count('id', filter=~Q(application_status='Created') &
                                            ~Q(application_status='Completed') &
                                            ~Q(application_status='Rejected') &
                                            ~Q(application_status='Fee Challan') &
                                            Q(assigned_group='DO')),
                PMC=Count('id', filter=~Q(application_status='Created') &
                                            ~Q(application_status='Completed') &
                                            ~Q(application_status='Rejected') &
                                            ~Q(application_status='Fee Challan') &
                                            ~Q(assigned_group='DO') &
                                            ~Q(assigned_group='APPLICANT')),
                APPLICANT=Count('id', filter=~Q(application_status='Created') &
                                        ~Q(application_status='Completed') &
                                        ~Q(application_status='Rejected') &
                                        ~Q(application_status='Fee Challan') &
                                        ~Q(assigned_group='DO') &
                                        Q(assigned_group='APPLICANT')),
                Licenses=Count('id', filter=Q(application_status='Completed')),
            ) \
            .order_by(
                Case(
                    When(registration_for='Producer', then=Value(1)),
                    When(registration_for='Consumer', then=Value(2)),
                    When(registration_for='Collector', then=Value(3)),
                    When(registration_for='Recycler', then=Value(4)),
                    default=Value(5), output_field=IntegerField()
                )  # Custom order
            )

        # Create a "Total" row by aggregating the counts
        total_statistics = registration_statistics.aggregate(
            total_applications=Sum('Applications'),
            total_do=Sum('DO'),
            total_pmc=Sum('PMC'),
            total_applicant=Sum('APPLICANT'),
            total_licenses=Sum('Licenses'),
        )

        total_row = {
            'registration_for': 'Total',
            'Applications': total_statistics['total_applications'],
            'DO': total_statistics['total_do'],
            'PMC': total_statistics['total_pmc'],
            'APPLICANT': total_statistics['total_applicant'],
            'Licenses': total_statistics['total_licenses']
        }
        
        registration_statistics = list(registration_statistics)
        registration_statistics.insert(0, total_row)


        # Grid data
        grid_data = ApplicantDetail.objects.values(
            'id',
            'first_name',
            'last_name',
            'cnic',
            'mobile_no',
            'application_status',
            'tracking_number',
            'assigned_group',
            'registration_for',
            'businessprofile__district__district_name'  # Include district name
        ).filter(~Q(application_status='Created') & ~Q(application_status='Fee Challan'))

        response_data = {
            # 'statistics': {
            #     'Applications': applications_count,
            #     'Licenses': licenses_count,
            #     'InProgress': in_progress_count
            # },
            'district_data': list(district_statistics),
            'grid_data': list(grid_data),
            'registration_statistics': list(registration_statistics),  # New key
        }

        return Response(response_data)

class MISApplicantStatisticsView(APIView):
    """
    API to calculate applicant statistics and return details for the grid.
    """
    permission_classes = []
    def get(self, request):
        district_statistics = ApplicantDetail.objects.exclude(
            application_status__in=['Created', 'Fee Challan']
        ) \
            .values('registration_for', 'businessprofile__district__district_name') \
            .annotate(count=Count('id')) \
            .order_by('businessprofile__district__district_name')

        
        # Statistics filtered by 'registration_for'
        registration_statistics = ApplicantDetail.objects \
            .exclude(registration_for__isnull=True) \
            .values('registration_for') \
            .annotate(
                Applications=Count('id', filter=~Q(application_status='Created') &
                                        ~Q(application_status='Fee Challan')),
                DO=Count('id', filter=~Q(application_status='Created') &
                                            ~Q(application_status='Completed') &
                                            ~Q(application_status='Rejected') &
                                            ~Q(application_status='Fee Challan') &
                                            Q(assigned_group='DO')),
                PMC=Count('id', filter=~Q(application_status='Created') &
                                            ~Q(application_status='Completed') &
                                            ~Q(application_status='Rejected') &
                                            ~Q(application_status='Fee Challan') &
                                            ~Q(assigned_group='DO') &
                                            ~Q(assigned_group='APPLICANT')),
                APPLICANT=Count('id', filter=~Q(application_status='Created') &
                                        ~Q(application_status='Completed') &
                                        ~Q(application_status='Rejected') &
                                        ~Q(application_status='Fee Challan') &
                                        ~Q(assigned_group='DO') &
                                        Q(assigned_group='APPLICANT')),
                Licenses=Count('id', filter=Q(application_status='Completed')),
            ) \
            .order_by(
                Case(
                    When(registration_for='Producer', then=Value(1)),
                    When(registration_for='Consumer', then=Value(2)),
                    When(registration_for='Collector', then=Value(3)),
                    When(registration_for='Recycler', then=Value(4)),
                    default=Value(5), output_field=IntegerField()
                )  # Custom order
            )

        # Create a "Total" row by aggregating the counts
        total_statistics = registration_statistics.aggregate(
            total_applications=Sum('Applications'),
            total_do=Sum('DO'),
            total_pmc=Sum('PMC'),
            total_applicant=Sum('APPLICANT'),
            total_licenses=Sum('Licenses'),
        )

        total_row = {
            'registration_for': 'Total',
            'Applications': total_statistics['total_applications'],
            'DO': total_statistics['total_do'],
            'PMC': total_statistics['total_pmc'],
            'APPLICANT': total_statistics['total_applicant'],
            'Licenses': total_statistics['total_licenses']
        }
        
        registration_statistics = list(registration_statistics)
        registration_statistics.insert(0, total_row)


        # Grid data
        grid_data = ApplicantDetail.objects.values(
            'id',
            'first_name',
            'last_name',
            'cnic',
            'mobile_no',
            'application_status',
            'tracking_number',
            'assigned_group',
            'registration_for',
            'businessprofile__district__district_name'  # Include district name
        ).filter(~Q(application_status='Created') & ~Q(application_status='Fee Challan'))

        response_data = {
            # 'statistics': {
            #     'Applications': applications_count,
            #     'Licenses': licenses_count,
            #     'InProgress': in_progress_count
            # },
            'district_data': list(district_statistics),
            'grid_data': list(grid_data),
            'registration_statistics': list(registration_statistics),  # New key
        }

        return Response(response_data)

def download_file(request, folder_name, file_name):
    # Construct the full file path dynamically
    file_path = os.path.join('/media/', folder_name, file_name)

    if os.path.exists(file_path):
        # Serve the file for inline viewing
        response = FileResponse(open(file_path, "rb"), as_attachment=False)
        response["Content-Disposition"] = f'inline; filename="{file_name}"'
        return response
    else:
        raise Http404("File does not exist")

def download_file2(request, folder_name, folder_name2, file_name):
    # Construct the full file path dynamically
    file_path = os.path.join('/media/', folder_name, folder_name2, file_name)

    if os.path.exists(file_path):
        # Serve the file for inline viewing
        response = FileResponse(open(file_path, "rb"), as_attachment=False)
        response["Content-Disposition"] = f'inline; filename="{file_name}"'
        return response
    else:
        raise Http404("File does not exist")

class ApplicantAlertsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user

        # 1) Get applicants for both "APPLICANT" and "Download License"
        applicants = ApplicantDetail.objects.filter(
            created_by=user,
            assigned_group__in=['APPLICANT', 'Download License']
        )

        # 2) Get IDs of those applicants
        applicant_ids = applicants.values_list('id', flat=True)

        # 3) Query assignments where assigned_group="APPLICANT" and has valid remarks
        assignments = ApplicationAssignment.objects.filter(
            applicant_id__in=applicant_ids,
            assigned_group='APPLICANT',
            remarks__isnull=False
        ).exclude(
            Q(remarks__isnull=True) | Q(remarks__iexact='undefined')
        ).order_by('-created_at')

        # 4) Serialize the original assignments
        serialized_data = ApplicantAlertsSerializer(assignments, many=True).data

        # 5) Get "Download License" applicants and manually add default remarks in response
        download_license_applicants = applicants.filter(assigned_group="Download License")
        for applicant in download_license_applicants:
            tracking_number = applicant.tracking_number if applicant.tracking_number else "N/A"
            serialized_data.append({
                "applicant_id": applicant.id,
                "assigned_group": "Download License",
                "remarks": f"Please Download License [{tracking_number}]",  # ✅ Tracking Number Added
                "url_sub_part": "/home-license"  # ✅ URL added
            })

        # 6) Return modified serialized data
        return Response(serialized_data, status=200)

# class ResetPasswordView(APIView):
#     def post(self, request):
#         current_password = request.data.get('current_password')
#         new_password = request.data.get('new_password')

#         user = request.user

#         if not user.check_password(current_password):
#             return Response({'detail': 'Current password is incorrect.'}, status=status.HTTP_400_BAD_REQUEST)

#         user.set_password(new_password)
#         user.save()
#         return Response({'message': 'Password updated successfully.'}, status=status.HTTP_200_OK)

class TrackApplicationView(APIView):
    permission_classes = []
    def get(self, request, *args, **kwargs):
        tracking_number = request.GET.get('tracking_number')
        if not tracking_number:
            return JsonResponse(
                {"message": "Tracking number is required. For further details, call helpline 1373."},
                status=400,
            )

        try:
            # Fetch application by tracking number
            application = ApplicantDetail.objects.get(tracking_number=tracking_number)
            assigned_group = application.assigned_group
            status_message = ""

            if not assigned_group or assigned_group.strip() == "":
                status_message = (
                    f"The application with Tracking Number '{tracking_number}' is in draft form. "
                    "Please complete it. For further details, call helpline 1373."
                )
            elif assigned_group in ['LSO', 'LSM', 'LSM2', 'TL']:
                status_message = (
                    f"The application with Tracking Number '{tracking_number}' is with the Plastic Management Cell "
                    "and is being processed. For further details, call helpline 1373."
                )
            elif assigned_group == 'DO':
                status_message = (
                    f"The application with Tracking Number '{tracking_number}' is with the Environment Officer District Incharge. "
                    "For further details, call helpline 1373."
                )
            elif assigned_group == 'APPLICANT':
                # Fetch the latest comment from ApplicationAssignment
                assignment = ApplicationAssignment.objects.filter(applicant=application).order_by('-created_at').first()
                comment = assignment.remarks if assignment and assignment.remarks != 'undefined' else "No reason provided."
                status_message = (
                    f"The application with Tracking Number '{tracking_number}' has been reassigned to the applicant. "
                    f"Reason: {comment}. For further details, call helpline 1373."
                )
            elif assigned_group == 'DEO':
                status_message = (
                    f"The application with Tracking Number '{tracking_number}' is with the Designated Environmental Officer. "
                    "For further details, call helpline 1373."
                )
            elif assigned_group == 'DG':
                status_message = (
                    f"The application with Tracking Number '{tracking_number}' is with the DG, EPA. "
                    "For further details, call helpline 1373."
                )
            elif assigned_group == 'Download License':
                status_message = (
                    f"The application with Tracking Number '{tracking_number}' has been processed and the license is ready for download. "
                    "You can download the license from your My Applications Dashboard. "
                    "For further details, call helpline 1373."
                )
            else:
                status_message = (
                    f"The application with Tracking Number '{tracking_number}' has an unknown status. "
                    "Please contact support. For further details, call helpline 1373."
                )

            return JsonResponse({"message": status_message}, status=200)

        except ApplicantDetail.DoesNotExist:
            return JsonResponse(
                {"message": f"No application found for the provided Tracking Number '{tracking_number}'. "
                            "For further details, call helpline 1373."},
                status=404,
            )

class ApplicantLocationViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = (
        ApplicantDetail.objects
        .select_related(
            'businessprofile__district',
            'businessprofile__tehsil',
            'manual_fields',
        )
        .filter(
            manual_fields__latitude__isnull=False,
            manual_fields__longitude__isnull=False
        )
    )
    serializer_class = ApplicantLocationSerializer
    permission_classes = []

class DistrictPlasticStatsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint that returns the district-wise plastic data.
    """
    permission_classes = []
    queryset = TblDistricts.objects.all()
    serializer_class = DistrictPlasticStatsSerializer

def sanitize_excel_row(row):
    sanitized = []
    for val in row.values():
        if isinstance(val, list):
            sanitized.append(", ".join(map(str, val)))
        elif val is None:
            sanitized.append("")
        else:
            sanitized.append(str(val))
    return sanitized

class InspectionReportViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = InspectionReportSerializer

    def get_queryset(self):
        """
        If the logged-in user has a district assigned, fetch only the reports from that district.
        Otherwise, fetch all reports.
        """
        user = self.request.user
        user_profile = UserProfile.objects.filter(user=user).first()  # Get user's profile

        if user_profile and user_profile.district:
            return InspectionReport.objects.filter(district=user_profile.district)
        
        return InspectionReport.objects.all()  # If no district is assigned, return all reports

    def perform_create(self, serializer):
        """
        Automatically assign the district from the user's profile when creating an inspection report.
        """
        user = self.request.user
        user_profile = UserProfile.objects.filter(user=user).first()

        if user_profile and user_profile.district:
            serializer.save(created_by=user, district=user_profile.district)
        else:
            serializer.save(created_by=user)  # Save without district if not found

    @action(detail=False, methods=['get'])
    def district_summary(self, request):
        """
        Fetch district-wise summary of inspection reports.
        """
        user = request.user
        user_profile = UserProfile.objects.filter(user=user).first()

        if user_profile and user_profile.district:
            summary = InspectionReport.objects.filter(district=user_profile.district)
        else:
            summary = InspectionReport.objects.all()

        summary = summary.values('district__district_name').annotate(
            total_inspections=Count('id'),
            total_notices_issued=Count('violation_found'),
            total_plastic_bags_confiscated=Sum('plastic_bags_confiscation', default=0),
            total_confiscated_plastic=Sum('total_confiscation', default=0),
            total_firs_registered=Count('action_taken', filter=Q(action_taken__contains='FIR')),
            total_premises_sealed=Count('action_taken', filter=Q(action_taken__contains='Sealed')),
            total_complaints_filed=Count('action_taken', filter=Q(action_taken__contains='Complaint')),

            # ✅ New KPIs Added
            total_fine_amount=Sum('fine_amount', default=0),
            total_fine_recovered=Sum('recovery_amount', default=0),
            pending_fine_amount=Sum('fine_amount', default=0) - Sum('recovery_amount', default=0),

            total_fines_pending=Count('fine_recovery_status', filter=Q(fine_recovery_status="Pending")),
            total_fines_partial=Count('fine_recovery_status', filter=Q(fine_recovery_status="Partial")),
            total_fines_recovered=Count('fine_recovery_status', filter=Q(fine_recovery_status="Recovered")),

            total_de_sealed_premises=Count('de_sealed_date', filter=Q(de_sealed_date__isnull=False)),
            total_affidavits_uploaded=Count('affidavit', filter=Q(affidavit__isnull=False)),
        )

        serializer = DistrictSummarySerializer(summary, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export-district-summary-excel')
    def export_district_summary_excel(self, request):
        user = request.user
        user_profile = getattr(user, 'userprofile', None)

        if user_profile and user_profile.district:
            queryset = InspectionReport.objects.filter(district=user_profile.district)
        else:
            queryset = InspectionReport.objects.all()

        summary = queryset.values('district__district_name').annotate(
            total_inspections=Count('id'),
            total_notices_issued=Count('violation_found'),
            total_plastic_bags_confiscated=Sum('plastic_bags_confiscation'),
            total_confiscated_plastic=Sum('total_confiscation'),
            total_firs_registered=Count('action_taken', filter=Q(action_taken__contains='FIR')),
            total_premises_sealed=Count('action_taken', filter=Q(action_taken__contains='Sealed')),
            total_complaints_filed=Count('action_taken', filter=Q(action_taken__contains='Complaint')),
            total_fine_amount=Sum('fine_amount'),
            total_fine_recovered=Sum('recovery_amount'),
            pending_fine_amount=Sum('fine_amount') - Sum('recovery_amount'),
            total_fines_pending=Count('fine_recovery_status', filter=Q(fine_recovery_status="Pending")),
            total_fines_partial=Count('fine_recovery_status', filter=Q(fine_recovery_status="Partial")),
            total_fines_recovered=Count('fine_recovery_status', filter=Q(fine_recovery_status="Recovered")),
            total_de_sealed_premises=Count('de_sealed_date', filter=Q(de_sealed_date__isnull=False)),
            total_affidavits_uploaded=Count('affidavit', filter=Q(affidavit__isnull=False)),
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "District Summary"

        # === Add PMIS logo ===
        logo_path = '/media/logo/logo-light-full.png'  # Adjust path if needed
        logo = OpenpyxlImage(logo_path)
        logo.width = 300
        logo.height = 100
        ws.add_image(logo, "A1")  # Add to top left

        # === Skip to Row 6 after logo
        for _ in range(5):
            ws.append([])

        # === Beautify headers
        raw_headers = list(summary[0].keys()) if summary else []
        beautified_headers = [
            header.replace('_', ' ')
                .replace('district__district_name', 'District Name')
                .title()
            for header in raw_headers
        ]
        ws.append(beautified_headers)

        # === Append data rows
        for row in summary:
            ws.append(list(row.values()))

        # === Style header row at row 6
        header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
        header_row_index = 6
        for col_index, cell in enumerate(ws[header_row_index], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

            col_letter = get_column_letter(col_index)
            ws.column_dimensions[col_letter].width = max(len(str(cell.value)) + 5, 20)

        # === Return as downloadable Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="district_summary.xlsx"'
        wb.save(response)
        return response


    @action(detail=False, methods=['get'], url_path='export-district-summary-pdf')
    def export_district_summary_pdf(self, request):
        user = request.user
        user_profile = getattr(user, 'userprofile', None)

        if user_profile and user_profile.district:
            queryset = InspectionReport.objects.filter(district=user_profile.district)
        else:
            queryset = InspectionReport.objects.all()

        summary = queryset.values('district__district_name').annotate(
            total_inspections=Count('id'),
            total_notices_issued=Count('violation_found'),
            total_plastic_bags_confiscated=Sum('plastic_bags_confiscation'),
            total_confiscated_plastic=Sum('total_confiscation'),
            total_firs_registered=Count('action_taken', filter=Q(action_taken__contains='FIR')),
            total_premises_sealed=Count('action_taken', filter=Q(action_taken__contains='Sealed')),
            total_complaints_filed=Count('action_taken', filter=Q(action_taken__contains='Complaint')),
            total_fine_amount=Sum('fine_amount'),
            total_fine_recovered=Sum('recovery_amount'),
            pending_fine_amount=Sum('fine_amount') - Sum('recovery_amount'),
            total_fines_pending=Count('fine_recovery_status', filter=Q(fine_recovery_status="Pending")),
            total_fines_partial=Count('fine_recovery_status', filter=Q(fine_recovery_status="Partial")),
            total_fines_recovered=Count('fine_recovery_status', filter=Q(fine_recovery_status="Recovered")),
            total_de_sealed_premises=Count('de_sealed_date', filter=Q(de_sealed_date__isnull=False)),
            total_affidavits_uploaded=Count('affidavit', filter=Q(affidavit__isnull=False)),
        )

        if summary:
            beautified_headers = [
                header.replace('_', ' ')
                    .replace('district__district_name', 'District Name')
                    .title()
                for header in list(summary[0].keys())
            ]
        else:
            beautified_headers = []

        # Prepare data for template
        context = {
            "title": "District-wise Inspection Summary",
            "headers": beautified_headers,
            "rows": [list(row.values()) for row in summary],
            "logo_path": request.build_absolute_uri("https://plmis.epapunjab.pk/img/logo/logo-light-full.png")
        }

        html = render_to_string("reports/district_summary_template.html", context)

        try:
            pdf = pdfkit.from_string(html, False, options={
                'page-size': 'A4',
                'encoding': "UTF-8",
                'enable-local-file-access': None  # required to access logo image
            })
        except Exception as e:
            return JsonResponse({"error": f"PDF generation failed: {str(e)}"}, status=500)

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="district_summary.pdf"'
        return response

    @action(detail=False, methods=['get'], url_path='export-all-inspections-excel')
    def export_all_inspections_excel(self, request):
        queryset = self.get_queryset().values(
            'id', 'district__district_name', 'violation_found', 'plastic_bags_confiscation',
            'total_confiscation', 'action_taken', 'fine_amount', 'recovery_amount',
            'fine_recovery_status', 'de_sealed_date', 'affidavit'
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "All Inspections"

        try:
            logo_path = '/media/logo/logo-light-full.png'
            logo = OpenpyxlImage(logo_path)
            logo.width = 300
            logo.height = 100
            ws.add_image(logo, "A1")
        except:
            pass

        for _ in range(5): ws.append([])

        headers = list(queryset[0].keys()) if queryset else []
        beautified_headers = [h.replace("_", " ").replace("district__district_name", "District").title() for h in headers]
        ws.append(beautified_headers)

        for row in queryset:
            ws.append(sanitize_excel_row(row))

        header_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        for col_index, cell in enumerate(ws[6], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_index)].width = max(len(str(cell.value)) + 5, 20)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="all_inspections.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export-all-inspections-pdf')
    def export_all_inspections_pdf(self, request):
        queryset = self.get_queryset().values(
            'id', 'district__district_name', 'violation_found', 'plastic_bags_confiscation',
            'total_confiscation', 'action_taken', 'fine_amount', 'recovery_amount',
            'fine_recovery_status', 'de_sealed_date', 'affidavit'
        )

        def sanitize(row):
            return [", ".join(map(str, v)) if isinstance(v, list) else str(v or "") for v in row.values()]

        headers = list(queryset[0].keys()) if queryset else []
        beautified_headers = [h.replace("_", " ").replace("district__district_name", "District").title() for h in headers]

        context = {
            "title": "All Inspection Reports",
            "headers": beautified_headers,
            "rows": [sanitize(row) for row in queryset],
            "logo_path": request.build_absolute_uri("https://plmis.epapunjab.pk/img/logo/logo-light-full.png")
        }

        html = render_to_string("reports/district_summary_template.html", context)

        try:
            pdf = pdfkit.from_string(html, False, options={
                'page-size': 'A4',
                'encoding': "UTF-8",
                'enable-local-file-access': None
            })
        except Exception as e:
            return JsonResponse({"error": f"PDF generation failed: {str(e)}"}, status=500)

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="all_inspections.pdf"'
        return response

    @action(detail=False, methods=['get'])
    def all_other_single_use_plastics(self, request):
        """Return the unique list of all single-use plastic items recorded"""
        snapshot, created = SingleUsePlasticsSnapshot.objects.get_or_create(id=1)
        return Response({"single_use_plastic_items": snapshot.plastic_items})
class DistrictPlasticCommitteeDocumentViewSet(viewsets.ModelViewSet):
    queryset = DistrictPlasticCommitteeDocument.objects.all()
    serializer_class = DistrictPlasticCommitteeDocumentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """
        If the logged-in user has a district assigned, fetch only the reports from that district.
        Otherwise, fetch all reports.
        """
        user = self.request.user
        user_profile = UserProfile.objects.filter(user=user).first()  # Get user's profile

        if user_profile and user_profile.district:
            return DistrictPlasticCommitteeDocument.objects.filter(district=user_profile.district)
        
        return DistrictPlasticCommitteeDocument.objects.all()  # If no district is assigned, return all reports
    
    def perform_create(self, serializer):
        user = self.request.user
        user_profile = UserProfile.objects.filter(user=user).first()  # Get the user's profile

        if user_profile and user_profile.district:
            serializer.save(uploaded_by=user, district=user_profile.district)
        else:
            raise serializers.ValidationError("User does not have an assigned district.")

class CompetitionRegistrationViewSet(viewsets.ModelViewSet):
    queryset = CompetitionRegistration.objects.all()
    serializer_class = CompetitionRegistrationSerializer
    parser_classes = [MultiPartParser, FormParser]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        registration = serializer.save()

        # Return only registration_id; React can use it to generate receipt
        return Response({
            "success": True,
            "registration_id": registration.registration_id
        }, status=status.HTTP_201_CREATED)

@require_GET
def generate_courier_label(request):
    registration_id = request.GET.get("registration_id")

    if not registration_id:
        return JsonResponse({"error": "registration_id is required"}, status=400)

    # Fetch the registered student info
    try:
        registration = get_object_or_404(CompetitionRegistration, registration_id=registration_id)
    except:
        return JsonResponse({"error": "Registration not found."}, status=404)

    # Prepare data for the label
    context = {
        "student_name": f"{registration.full_name or ''}",
        "class_name": registration.grade or "N/A",
        "institution": registration.institute or "N/A",
        "registration_id": registration.registration_id,
        "mobile": registration.mobile,
        "competition_name": registration.competition_type,
    }

    # Render HTML to PDF
    html = render_to_string("pdf_templates/courier_label_template.html", context)

    try:
        pdf = pdfkit.from_string(html, False, options={
            'page-size': 'A5',
            'encoding': "UTF-8",
            'enable-local-file-access': None
        })
    except Exception as e:
        return JsonResponse({"error": f"PDF generation failed: {str(e)}"}, status=500)

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="courier_label_{registration.registration_id}.pdf"'
    return response


def confiscation_lookup(request):
    book_no = request.GET.get('book_no')
    receipt_no = request.GET.get('receipt_no')

    try:
        report = InspectionReport.objects.get(
            receipt_book_number=book_no,
            receipt_number=receipt_no
        )
        return JsonResponse({
            "total_confiscation": report.total_confiscation,
            "receipt_book_number": report.receipt_book_number,
            "receipt_number": report.receipt_number,
            "receipt_url": request.build_absolute_uri(report.confiscation_receipt.url) if report.confiscation_receipt else None
        })
    except InspectionReport.DoesNotExist:
        return JsonResponse({"error": "Record not found."}, status=404)